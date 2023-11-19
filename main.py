import os
import logging
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Text, Command
from aiogram.utils import executor
from openpyxl import Workbook, load_workbook
from datetime import datetime
from sqlalchemy import func
from aiogram.types import CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton

from handlers import markups as nav
from config import class_for_main1 as nav2


# ---- ADMIN ID ----

admin_tg_id = [1238343405, 5459503530]


if not os.path.exists('data'):
    os.makedirs('data', exist_ok=True)

if not os.path.exists('data/production.xlsx'):
    open('data/production.xlsx', 'w').close()

if not os.path.exists('data/consumption.xlsx'):
    open('data/consumption.xlsx', 'w').close()


dp = nav2.dp

logging.basicConfig(level=logging.INFO)


async def is_admin_user(message: types.Message):
    return message.from_user.id in admin_tg_id


@dp.message_handler(Command("start"), is_admin_user)
async def cmd_start(message: types.Message):
    await message.reply("Привет! Я бот для учета производства. Нажмите кнопку \"Начать\", чтобы начать работу.",
                        reply_markup=nav.mainMenu)


# --------- SEND BUG REPORT ---------

@dp.message_handler(Text(equals="Отправить баг 🐛", ignore_case=True))
async def bug_report_command(message: types.Message):
    await message.reply("Чтобы отправить баг-репорт, прикрепите скриншот или опишите проблему.")
    await nav2.FormBugReport.text.set()


@dp.message_handler(state=nav2.FormBugReport.text, content_types=types.ContentType.ANY)
async def process_bug_report(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    chat_id = '-1002127891568'

    try:
        caption = f"Баг-репорт от пользователя {user_id}:\n\n{message.text}"
        await nav2.bot.send_message(chat_id, caption)

        if message.photo:
            photo = message.photo[-1].file_id
            await nav2.bot.send_photo(chat_id, photo, caption=caption)

        await message.reply("Баг-репорт успешно отправлен. Спасибо!")
    except Exception as e:
        logging.error(f"Error processing bug report: {e}")
        await message.reply("Произошла ошибка при обработке баг-репорта. Попробуйте позже.")

    await state.finish()


# --------- SUPPORT ---------

@dp.message_handler(Text(equals="Техподдержка 🛠", ignore_case=True))
async def support_report_command(message: types.Message):
    await message.reply(
        "🤖 <b>Швейный Учетный Бот - Техподдержка</b>\n\n"
        "Если у Вас возникли вопросы, предложения или вам нужна помощь с использованием бота, "
        "не стесняйтесь обращаться! Мы готовы ответить на все ваши запросы и обеспечить качественную поддержку.\n\n"
        "📱 <b>Телеграм:</b> <b>@al1shka007</b>\n\n"
        "Не забывайте указывать Ваши контактные данные и описывать проблему максимально подробно. "
        "Мы стремимся сделать использование бота максимально удобным для Вас!\n\n"
        "<b>Спасибо за выбор Швейного Учетного Бота!</b> 🚀",
        parse_mode='HTML'
    )


# --------- REPORTS ---------

@dp.message_handler(Text(equals="Отчет 📊", ignore_case=True))
async def process_name(message: types.Message):
    await message.reply("Введите ФИО мастера:")
    await nav2.FormReports.name.set()


@dp.message_handler(state=nav2.FormReports.name)
async def process_model_name(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['name'] = message.text
    await message.reply("Введите название модели:")
    await nav2.FormReports.model_name.set()


@dp.message_handler(state=nav2.FormReports.model_name)
async def process_model_name(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['model_name'] = message.text
    await message.reply("Введите количество изделий:")
    await nav2.FormReports.remaining.set()


@dp.message_handler(state=nav2.FormReports.remaining)
async def process_remaining(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['remaining'] = message.text
    await message.reply("Введите кол-во сколько приняли изделия:")
    await nav2.FormReports.income.set()


@dp.message_handler(state=nav2.FormReports.income)
async def process_unit_price(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['income'] = message.text
    await message.reply("Введите цену за единицу:")
    await nav2.FormReports.expenses.set()


# --------- SEND REPORTS FILE ---------

@dp.message_handler(state=nav2.FormReports.expenses)
async def process_reports(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['expenses'] = int(message.text)
        data['result_reports'] = str(int(data['income']) * int(data['expenses']))  # Итого = income * expenses

        new_report = nav2.Report(
            name=data['name'],
            model_name=data['model_name'],
            remaining=data['remaining'],
            income=data['income'],
            expenses=data['expenses'],
            result_reports=data['result_reports']
        )
        nav2.session.add(new_report)
        nav2.session.commit()

        try:
            file_path_reports = 'data/production.xlsx'
            if os.path.exists(file_path_reports):
                wb = load_workbook(file_path_reports)
            else:
                wb = load_workbook(file_path_reports)
                ws = wb.active
                ws.append(["Дата", "Название модели", "Мастер ФИО", "Количество", "Принял", "Цена", "Итого"])

            ws = wb.active
            income = int(data['income'])
            remaining = int(data['remaining'])
            expenses = int(data['expenses'])
            result_reports = int(data['result_reports'])
            model_name = str(data['model_name'])
            name = str(data['name'])
            row = (datetime.now().strftime("%d.%m.%Y"), model_name, name, income, remaining, expenses, result_reports)
            ws.append(row)
            wb.save(file_path_reports)

            with open(file_path_reports, 'rb') as f:
                await nav2.bot.send_document(message.chat.id, f)
        except Exception as e:
            logging.error(f"Error sending document: {e}")
            print(f"Error sending document: {e}")

    await state.finish()


# --------- SEARCH IN REPORTS BY NAME ---------

@dp.message_handler(Text(equals="Поиск 🔍", ignore_case=True))
async def search_options(message: types.Message):
    print("Получено сообщение 'Поиск'")
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("По имени 🕵️‍♂️", callback_data="search_by_name"))
    keyboard.add(InlineKeyboardButton("По названию модели 📄", callback_data="search_by_model"))

    await message.reply("Выберите опцию поиска:", reply_markup=keyboard)


@dp.callback_query_handler(lambda c: c.data.startswith("search"))
async def search_option_callback(callback_query: CallbackQuery):
    option = callback_query.data[len("search_"):]
    if option == "by_name":
        logging.info("Received request to search by name")
        await nav2.bot.send_message(callback_query.from_user.id, "Введите имя мастера для поиска:")
        current_state = await nav2.dp.storage.get_state(chat=callback_query.message.chat.id,
                                                        user=callback_query.from_user.id)
        logging.info("Current state: %s", current_state)
        await nav2.FormSearch.name_db.set()
    elif option == "by_model":
        logging.info("Received request to search by model")
        await nav2.bot.send_message(callback_query.from_user.id, "Введите название модели для поиска:")
        current_state = await nav2.dp.storage.get_state(chat=callback_query.message.chat.id,
                                                        user=callback_query.from_user.id)
        logging.info("Current state: %s", current_state)
        await nav2.FormSearch.model_db.set()


@dp.message_handler(state=nav2.FormSearch.name_db)
async def process_search_by_name(message: types.Message, state: FSMContext):
    logging.info("Received message for name search: %s", message.text)
    async with state.proxy() as data:
        master_name = message.text.lower()

        try:
            results = nav2.session.query(nav2.Report).filter(func.lower(nav2.Report.name) == master_name).all()

            if not results:
                await message.reply(f"Нет результатов поиска для мастера '{master_name}'.")
                await state.finish()
                return

            file_path = "data/search_results_name.xlsx"

            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active
            except FileNotFoundError:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Мастер", "Модель", "Количество", "Принято", "Итог"])

            for result in results:
                sheet.append([result.name, result.model_name, result.remaining, result.income, result.result_reports])

            workbook.save(file_path)
            workbook.close()

            with open(file_path, 'rb') as file:
                caption = f"Результаты поиска по мастеру '{master_name}'"
                await message.answer_document(file, caption=caption)

        except Exception as e:
            logging.error("Error querying the database: %s", str(e))
            await message.reply("Произошла ошибка при выполнении запроса к базе данных.")

    await state.finish()


# --------- SEARCH IN REPORTS BY MODEL NAME ---------

@dp.message_handler(state=nav2.FormSearch.model_db)
async def process_search_by_model(message: types.Message, state: FSMContext):
    logging.info("Received message for model name search: %s", message.text)
    async with state.proxy() as data:
        model_name = message.text.lower()

        try:
            results = nav2.session.query(nav2.Report).filter(func.lower(nav2.Report.model_name) == model_name).all()

            if not results:
                await message.reply(f"Нет результатов поиска для модели '{model_name}'.")
                await state.finish()
                return

            file_path = f"data/search_results_model.xlsx"

            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active
            except FileNotFoundError:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Мастер", "Модель", "Количество", "Принято", "Итог"])

            for result in results:
                sheet.append([result.name, result.model_name, result.remaining, result.income, result.result_reports])

            workbook.save(file_path)
            workbook.close()

            with open(file_path, 'rb') as file:
                caption = f"Результаты поиска по модели '{model_name}'"
                await message.answer_document(file, caption=caption)
        except Exception as e:
            logging.error("Error querying the database: %s", str(e))
            await message.reply("Произошла ошибка при выполнении запроса к базе данных.")

    await state.finish()


# --------- EXPENSES ---------

@dp.message_handler(Text(equals="Расходы 💸", ignore_case=True))
async def cmd_calc(message: types.Message):
    await message.reply("Введите расход за ткань:")
    await nav2.FormExpenses.textile.set()


@dp.message_handler(state=nav2.FormExpenses.textile)
async def expenses_by_accessories(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['textile'] = message.text
    await message.reply("Введите расходы за фурнитуру:")
    await nav2.FormExpenses.accessories.set()


@dp.message_handler(state=nav2.FormExpenses.accessories)
async def expenses_by_sewing(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['accessories'] = message.text
    await message.reply("Введите расходы на пошив за единицу:")
    await nav2.FormExpenses.sewing.set()


# --------- SEND EXPENSES FILE ---------

@dp.message_handler(state=nav2.FormExpenses.sewing)
async def process_expenses(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['sewing'] = message.text
        data['result_expenses'] = str(int(data['textile']) + int(data['accessories']) + int(data['sewing']))

        new_expenses = nav2.Expenses(
            textile=data['textile'],
            accessories=data['accessories'],
            sewing=data['sewing'],
            result_expenses=data['result_expenses']
        )
        nav2.session.add(new_expenses)
        nav2.session.commit()

        try:
            file_path_expenses = 'data/consumption.xlsx'
            if os.path.exists(file_path_expenses):
                wb = load_workbook(file_path_expenses)
            else:
                wb = load_workbook(file_path_expenses)
                ws = wb.active
                ws.append(['Дата', 'Закуп ткани', 'Фурнитура', 'Пошив за ед.', 'Итого'])

            ws = wb.active
            textile = int(data['textile'])
            accessories = int(data['accessories'])
            sewing = int(data['sewing'])
            result_expenses = int(data['result_expenses'])
            row = (datetime.now().strftime("%d.%m.%Y"), textile, accessories, sewing, result_expenses)
            ws.append(row)
            wb.save(file_path_expenses)

            with open(file_path_expenses, 'rb') as f:
                await nav2.bot.send_document(message.chat.id, f)
        except Exception as e:
            logging.error(f"Error sending document: {e}")
            print(f"Error sending document: {e}")

    await state.finish()


if __name__ == '__main__':
    executor.start_polling(nav2.dp, skip_updates=True)
